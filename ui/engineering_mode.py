# Sensor_Testor/ui/engineering_mode.py
from __future__ import annotations

import os
import csv
import copy
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

from PyQt5.QtCore import Qt, QThread
from PyQt5.QtWidgets import (
        QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout, QGridLayout,
        QLabel, QLineEdit, QPushButton, QFileDialog, QTextEdit, QMessageBox,
        QProgressBar, QCheckBox, QComboBox, QTableWidget, QTableWidgetItem,
        QHeaderView, QTabWidget, QInputDialog, QTableWidgetItem, QTableWidget, QDialogButtonBox
)

import pyqtgraph as pg
import math
from openpyxl import load_workbook


# -------------------------
# Robust imports (match your repo layout)
# -------------------------

# store + RunConfig
try:
        from Sensor_Testor.domain.models import store, RunConfig
except Exception:
        try:
                from domain.models import store, RunConfig  # type: ignore
        except Exception:
                from models import store, RunConfig  # type: ignore

# strict plan parser -> TestStep list + settings dict
try:
        from Sensor_Testor.app_io.plan_loader import load_grid_plan_csv
except Exception:
        try:
                from app_io.plan_loader import load_grid_plan_csv  # type: ignore
        except Exception as e:
                raise ImportError(f"Could not import load_grid_plan_csv from plan_loader: {e}")

# writers (factory)
try:
        from Sensor_Testor.app_io.writers import make_writers
except Exception:
        from app_io.writers import make_writers  # type: ignore

# criteria helper (optional)
try:
        from Sensor_Testor.processing.criteria_loader import (
                parse_pass_fail_criteria_form, generate_smoothed_line
        )
except Exception:
        try:
                from processing.criteria_loader import (
                        parse_pass_fail_criteria_form, generate_smoothed_line
                )
        except Exception:
                parse_pass_fail_criteria_form = None  # type: ignore

                def generate_smoothed_line(x, y, *_args, **_kwargs):  # type: ignore
                        return x, y

# adapters
try:
        from Sensor_Testor.hardware.duet_adapter import DuetAdapter
except Exception:
        try:
                from hardware.duet_adapter import DuetAdapter  # type: ignore
        except Exception:
                from duet_adapter import DuetAdapter  # type: ignore

try:
        from Sensor_Testor.hardware.daq_adapter import DaqAdapter
except Exception:
        try:
                from hardware.daq_adapter import DaqAdapter  # type: ignore
        except Exception:
                from daq_adapter import DaqAdapter  # type: ignore

try:
        from Sensor_Testor.hardware.smac_adapter import SmacAdapter
except Exception:
        try:
                from hardware.smac_adapter import SmacAdapter  # type: ignore
        except Exception:
                SmacAdapter = None  # type: ignore

# test runner worker (+ position-only worker used by the "Test Positions" button)
try:
        from Sensor_Testor.runner.test_runner import TestRunnerWorker, PositionOnlyWorker
except Exception:
        try:
                from runner.test_runner import TestRunnerWorker, PositionOnlyWorker  # type: ignore
        except Exception:
                from test_runner import TestRunnerWorker, PositionOnlyWorker  # type: ignore

# grid runner — drives XY travel / safe-Z clearance across all plan rows;
# used by "Test Positions" to reuse the exact same motion logic as a real run
try:
        from Sensor_Testor.runner.grid_runner import GridRunner
except Exception:
        try:
                from runner.grid_runner import GridRunner  # type: ignore
        except Exception:
                from grid_runner import GridRunner  # type: ignore


# -------------------------
# Small helpers
# -------------------------

def _to_bool(v: Any) -> bool:
        if isinstance(v, bool):
                return v
        s = str(v).strip().lower()
        return s in ("true", "1", "yes", "y", "on")

def _safe_float(s: str, default: float = 0.0) -> float:
        try:
                return float(str(s).strip())
        except Exception:
                return default

def _safe_int(s: str, default: int = 0) -> int:
        try:
                return int(float(str(s).strip()))
        except Exception:
                return default

def _header_index(table: QTableWidget, header_name: str) -> int:
        for i in range(table.columnCount()):
                it = table.horizontalHeaderItem(i)
                if it and it.text().strip().lower() == header_name.strip().lower():
                        return i
        return -1


# -------------------------
# Pass/Fail criteria builder (copied conceptually from old GUI.py, cleaned up)
# -------------------------

@dataclass
class CriteriaRow:
        type: str               # "at specific force" or "at specific resistance"
        value: Optional[float]  # force or resistance (depends on type)
        max: Optional[float]
        min: Optional[float]
        other: str = ""


class PassFailCriteriaDialog(QDialog):
        """
        Compact dialog to build a criteria file and store it into EngineeringMode.pending_criteria_files
        so it gets written alongside the project on save.

        This is intentionally simpler than the old UI: we keep the same data model
        (Type / Value / Max / Min / Other), but avoid duplicated widgets and fragile state.
        """

        def __init__(self, parent=None, engineering_mode=None, apply_row: Optional[int] = None):
                super().__init__(parent)
                self.setWindowTitle("Pass/Fail Criteria")
                self.setModal(True)
                self.setMinimumSize(820, 520)

                self.engineering_mode = engineering_mode
                self.apply_row = apply_row

                self.criteria_rows: List[CriteriaRow] = []

                root = QVBoxLayout(self)

                # --- Top entry area (two compact groups) ---
                top = QHBoxLayout()

                g_res = QGroupBox("Resistance limits at a force", self)
                fr = QFormLayout(g_res)
                self.ed_force_for_res = QLineEdit(self)
                self.ed_max_res = QLineEdit(self)
                self.ed_min_res = QLineEdit(self)
                fr.addRow("Force (N):", self.ed_force_for_res)
                fr.addRow("Max Resistance (Ω):", self.ed_max_res)
                fr.addRow("Min Resistance (Ω):", self.ed_min_res)
                top.addWidget(g_res, 1)

                g_force = QGroupBox("Force limits at a resistance", self)
                ff = QFormLayout(g_force)
                self.ed_res_for_force = QLineEdit(self)
                self.ed_max_force = QLineEdit(self)
                self.ed_min_force = QLineEdit(self)
                ff.addRow("Resistance (Ω):", self.ed_res_for_force)
                ff.addRow("Max Force (N):", self.ed_max_force)
                ff.addRow("Min Force (N):", self.ed_min_force)
                top.addWidget(g_force, 1)

                root.addLayout(top)

                btn_row = QHBoxLayout()
                self.btn_add = QPushButton("Add point", self)
                self.btn_add.clicked.connect(self._add_point)
                self.btn_clear = QPushButton("Clear", self)
                self.btn_clear.clicked.connect(self._clear_points)
                self.btn_save = QPushButton("Save to project…", self)
                self.btn_save.clicked.connect(self._save_into_project)

                btn_row.addWidget(self.btn_add)
                btn_row.addWidget(self.btn_clear)
                btn_row.addStretch(1)
                btn_row.addWidget(self.btn_save)
                root.addLayout(btn_row)

                # --- Table + plot preview ---
                mid = QHBoxLayout()
                self.table = QTableWidget(self)
                self.table.setColumnCount(5)
                self.table.setHorizontalHeaderLabels(["Type", "Value", "Max", "Min", "Other"])
                self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                mid.addWidget(self.table, 2)

                self.plot = pg.PlotWidget(title="Criteria preview (Force vs Resistance)")
                self.plot.setLabel("bottom", "Force (N)")
                self.plot.setLabel("left", "Resistance (Ω)")
                mid.addWidget(self.plot, 3)

                root.addLayout(mid, 1)

                info = QLabel(
                        "Tip: Add either (Force + Max/Min Resistance) OR (Resistance + Max/Min Force). "
                        "You can combine multiple points; they will be saved as a CSV criteria file.",
                        self
                )
                info.setWordWrap(True)
                root.addWidget(info)

        def _add_point(self):
                added = False

                # Case A: limits on resistance at a force
                if self.ed_force_for_res.text().strip() and (self.ed_max_res.text().strip() or self.ed_min_res.text().strip()):
                        f = _safe_float(self.ed_force_for_res.text(), None)  # type: ignore[arg-type]
                        if f is None:
                                QMessageBox.warning(self, "Input Error", "Force must be a number.")
                                return
                        mx = self.ed_max_res.text().strip()
                        mn = self.ed_min_res.text().strip()
                        row = CriteriaRow(
                                type="at specific force",
                                value=float(f),
                                max=_safe_float(mx, None) if mx else None,  # type: ignore[arg-type]
                                min=_safe_float(mn, None) if mn else None,  # type: ignore[arg-type]
                                other=""
                        )
                        self.criteria_rows.append(row)
                        added = True

                # Case B: limits on force at a resistance
                if self.ed_res_for_force.text().strip() and (self.ed_max_force.text().strip() or self.ed_min_force.text().strip()):
                        r = _safe_float(self.ed_res_for_force.text(), None)  # type: ignore[arg-type]
                        if r is None:
                                QMessageBox.warning(self, "Input Error", "Resistance must be a number.")
                                return
                        mx = self.ed_max_force.text().strip()
                        mn = self.ed_min_force.text().strip()
                        row = CriteriaRow(
                                type="at specific resistance",
                                value=float(r),
                                max=_safe_float(mx, None) if mx else None,  # type: ignore[arg-type]
                                min=_safe_float(mn, None) if mn else None,  # type: ignore[arg-type]
                                other=""
                        )
                        self.criteria_rows.append(row)
                        added = True

                if not added:
                        QMessageBox.warning(self, "Input Error", "Enter either Force + Max/Min Resistance OR Resistance + Max/Min Force.")
                        return

                self._sync_table()
                self._sync_plot()

        def _clear_points(self):
                self.criteria_rows = []
                self.table.setRowCount(0)
                self.plot.clear()

        def _sync_table(self):
                self.table.setRowCount(0)
                for r in self.criteria_rows:
                        i = self.table.rowCount()
                        self.table.insertRow(i)
                        self.table.setItem(i, 0, QTableWidgetItem(r.type))
                        self.table.setItem(i, 1, QTableWidgetItem("" if r.value is None else str(r.value)))
                        self.table.setItem(i, 2, QTableWidgetItem("" if r.max is None else str(r.max)))
                        self.table.setItem(i, 3, QTableWidgetItem("" if r.min is None else str(r.min)))
                        self.table.setItem(i, 4, QTableWidgetItem(r.other or ""))

        def _sync_plot(self):
                """
                Plot two curves:
                  - MAX curve (dashed)
                  - MIN curve (dashed)
                If smoothing helper exists, smooth them.
                """
                self.plot.clear()

                # Collect points into (force, res) pairs.
                max_pts: List[tuple[float, float]] = []
                min_pts: List[tuple[float, float]] = []

                for r in self.criteria_rows:
                        if r.type == "at specific force" and r.value is not None:
                                force_n = float(r.value)
                                if r.max is not None:
                                        max_pts.append((force_n, float(r.max)))
                                if r.min is not None:
                                        min_pts.append((force_n, float(r.min)))
                        elif r.type == "at specific resistance" and r.value is not None:
                                # Here "value" is resistance, and max/min are forces.
                                res_ohm = float(r.value)
                                if r.max is not None:
                                        max_pts.append((float(r.max), res_ohm))
                                if r.min is not None:
                                        min_pts.append((float(r.min), res_ohm))

                def _plot(points: List[tuple[float, float]], which: str):
                        if len(points) < 2:
                                # If only one point, still show it.
                                if len(points) == 1:
                                        x, y = points[0]
                                        self.plot.plot([x], [y], symbol="o")
                                return

                        points = sorted(set(points), key=lambda p: p[0])
                        xs = [p[0] for p in points]
                        ys = [p[1] for p in points]

                        try:
                                xs_s, ys_s = generate_smoothed_line(xs, ys, 200)
                        except Exception:
                                xs_s, ys_s = xs, ys

                        # Let pyqtgraph pick default pen; we avoid hardcoded colors.
                        self.plot.plot(xs_s, ys_s)

                _plot(max_pts, "max")
                _plot(min_pts, "min")

                # ensure some auto range
                try:
                        vb = self.plot.getViewBox()
                        vb.enableAutoRange(axis=pg.ViewBox.XAxis, enable=True)
                        vb.enableAutoRange(axis=pg.ViewBox.YAxis, enable=True)
                except Exception:
                        self.plot.enableAutoRange(True, True)

        def _save_into_project(self):
                if not self.criteria_rows:
                        QMessageBox.warning(self, "No data", "Add at least one criteria point before saving.")
                        return

                name, ok = QInputDialog.getText(
                        self, "Criteria file name", "Enter a name (no extension):", text="criteria"
                )
                if not ok or not name.strip():
                        return

                filename = name.strip() + ".csv"
                payload = [r.__dict__ for r in self.criteria_rows]

                if self.engineering_mode is None:
                        QMessageBox.warning(self, "Error", "No EngineeringMode context attached.")
                        return

                self.engineering_mode.pending_criteria_files[filename] = copy.deepcopy(payload)

                # Apply criteria file into the table:
                try:
                        self.engineering_mode.apply_pf_filename(filename, only_row=self.apply_row)
                except Exception:
                        pass

                QMessageBox.information(self, "Saved", f"Criteria saved as '{filename}' (will be written on project save).")
                self.accept()

class GoldenCurveDialog(QDialog):
        """
        Popup for loading an XLSX and plotting all *Filtered* datasets found in it.

        - User selects .xlsx
        - Presses Generate
        - Dialog parses all sheets containing 'Filtered' in the sheet name
          (fallback: all sheets if none match)
        - It auto-detects Force + Resistance columns by header text.
        """

        def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Golden Curve")
                self.resize(900, 650)

                self.xlsx_path: str | None = None

                root = QVBoxLayout(self)

                # --- file row
                file_row = QHBoxLayout()
                self.ed_xlsx = QLineEdit(self)
                self.ed_xlsx.setPlaceholderText("Select an .xlsx containing filtered data...")
                btn_browse = QPushButton("Browse", self)
                btn_browse.clicked.connect(self._browse_xlsx)
                file_row.addWidget(self.ed_xlsx, 1)
                file_row.addWidget(btn_browse)
                root.addLayout(file_row)

                # --- plot
                self.plot = pg.PlotWidget(title="Golden Curve – Filtered Data")
                self.plot.showGrid(x=True, y=True)
                self.plot.setLabel("bottom", "Force")
                self.plot.setLabel("left", "Resistance")
                root.addWidget(self.plot, 1)

                # --- controls row (Generate + Close)
                controls = QHBoxLayout()
                controls.addStretch(1)

                self.btn_generate = QPushButton("Generate", self)
                self.btn_generate.clicked.connect(self._generate)
                controls.addWidget(self.btn_generate)

                bb = QDialogButtonBox(QDialogButtonBox.Close, self)
                bb.rejected.connect(self.reject)
                controls.addWidget(bb)

                root.addLayout(controls)

        def _browse_xlsx(self):
                p, _ = QFileDialog.getOpenFileName(self, "Select XLSX", "", "Excel Files (*.xlsx);;All Files (*)")
                if p:
                        self.xlsx_path = p
                        self.ed_xlsx.setText(p)

        def _generate(self):
                try:
                        path = (self.xlsx_path or "").strip()
                        if not path:
                                path = self.ed_xlsx.text().strip()
                        if not path:
                                raise ValueError("Please select an .xlsx file first.")
                        if not os.path.exists(path):
                                raise FileNotFoundError(f"File not found:\n{path}")

                        wb = load_workbook(path, data_only=True)

                        # Prefer sheets containing "Filtered"
                        filtered_sheets = [s for s in wb.sheetnames if "filtered" in s.lower()]
                        target_sheets = filtered_sheets if filtered_sheets else wb.sheetnames

                        self.plot.clear()

                        any_plotted = False
                        for sheet_name in target_sheets:
                                ws = wb[sheet_name]

                                # Read first row as headers
                                headers = []
                                for c in range(1, ws.max_column + 1):
                                        v = ws.cell(row=1, column=c).value
                                        headers.append(str(v).strip() if v is not None else "")

                                # Find likely columns for Force and Resistance
                                def find_col(possible_words):
                                        for idx, h in enumerate(headers, start=1):
                                                hl = h.lower()
                                                if any(w in hl for w in possible_words):
                                                        return idx
                                        return None

                                # You can expand these keywords if your headers differ
                                force_col = find_col(["force", "load"])
                                res_col = find_col(["resistance", "ohm", "r (", "res ("])

                                # If not found by headers, skip this sheet
                                if force_col is None or res_col is None:
                                        continue

                                x = []
                                y = []
                                for r in range(2, ws.max_row + 1):
                                        fx = ws.cell(row=r, column=force_col).value
                                        ry = ws.cell(row=r, column=res_col).value

                                        # Skip blanks / non-numerics
                                        try:
                                                fx = float(fx)
                                                ry = float(ry)
                                                if math.isnan(fx) or math.isnan(ry):
                                                        continue
                                        except Exception:
                                                continue

                                        x.append(fx)
                                        y.append(ry)

                                if len(x) >= 2:
                                        # Plot this sheet as one curve
                                        self.plot.plot(x, y, name=sheet_name)
                                        any_plotted = True

                        if not any_plotted:
                                raise ValueError(
                                        "No filtered datasets were found.\n\n"
                                        "Make sure your XLSX has sheets with 'Filtered' in the name, "
                                        "and columns with headers like 'Force' and 'Resistance'."
                                )

                except Exception as e:
                        QMessageBox.critical(self, "Golden Curve Error", str(e))
# -------------------------
# Engineering Mode
# -------------------------

class EngineeringMode(QDialog):
        """
        Engineering Mode supports:
          1) Build/Save/Load plan projects (Grid plan CSV + criteria files folder)
          2) Run a plan directly with the new modular runner (TestRunnerWorker)

        This is a refactor of the old GUI.py EngineeringMode into the new ui/engineering_mode.py,
        keeping behavior but removing duplication and global state.
        """

        TABLE_HEADERS = [
                "Test", "X Position", "Y Position", "Force",
                "Speed of Test", "Speed between Test",
                "Safe Height (mm)", "Test Height (mm)",
                "Golden Curve", "P/F Criteria",
        ]

        def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Sensor Tester – Engineering Mode")
                self.resize(1300, 860)

                # project state
                self.project_folder: Optional[str] = None
                self.pass_fail_folder: Optional[str] = None
                self.grid_test_csv_path: Optional[str] = None

                # criteria staged in-memory until project save
                self.pending_criteria_files: Dict[str, List[Dict[str, Any]]] = {}

                # hardware
                self.duet = DuetAdapter()
                self.smac = SmacAdapter() if SmacAdapter is not None else None

                self._daq_ok = False
                self.daq = None
                try:
                        self.daq = DaqAdapter(channels=(0, 1, 2, 3), rate_hz=1000.0)
                        if hasattr(self.daq, "open"):
                                self.daq.open()
                        self._daq_ok = True
                except Exception:
                        self.daq = None
                        self._daq_ok = False

                # run worker
                self.thread: Optional[QThread] = None
                self.worker: Optional[TestRunnerWorker] = None

                # "Test Positions" dry-run worker (motion only, no DAQ/probe)
                self.position_thread: Optional[QThread] = None
                self.position_worker: Optional[PositionOnlyWorker] = None
                self.position_runner: Optional[GridRunner] = None

                # thresholds stored when those checkboxes are enabled
                self.preload_threshold_ohm: Optional[float] = None
                self.shortcircuit_threshold_ohm: Optional[float] = None

                # UI
                self._build_ui()

        # ---------- UI ----------
        def _build_ui(self):
                root = QVBoxLayout(self)

                self.tabs = QTabWidget(self)
                root.addWidget(self.tabs, 1)

                # --- Build tab ---
                self.tab_build = QDialog(self)
                self.tabs.addTab(self.tab_build, "Build / Edit Plan")
                self._build_tab_build()

                # --- Run tab ---
                self.tab_run = QDialog(self)
                self.tabs.addTab(self.tab_run, "Run Plan")
                self._build_tab_run()

        def _build_tab_build(self):
                outer = QHBoxLayout(self.tab_build)

                # left panel: settings + actions
                left = QVBoxLayout()
                outer.addLayout(left, 2)

                # --- Test type & starting position ---
                g_meta = QGroupBox("Plan Basics", self.tab_build)
                meta = QFormLayout(g_meta)

                self.dd_test_type = QComboBox(self.tab_build)
                self.dd_test_type.addItems(["Grid Test", "Location Test", "Simple Test"])

                self.dd_starting_position = QComboBox(self.tab_build)
                self.dd_starting_position.addItems(["Top Right", "Top Left", "Bottom Right", "Bottom Left"])

                meta.addRow("Test Type:", self.dd_test_type)
                meta.addRow("Starting Position:", self.dd_starting_position)
                left.addWidget(g_meta)

                # --- Settings entries (aligned like old) ---
                g_settings = QGroupBox("Settings", self.tab_build)
                grid = QGridLayout(g_settings)

                self.setting_names = [
                        "x pitch(mm)", "y pitch(mm)",
                        "Number of Sensors in x", "Number of Sensors in y",
                        "actuator speed(mm/s)", "speed between spaces(mm/s)",
                        "max force(kg)", "start force(kg)",
                        "start position x", "start position y",
                        "Safe Height (mm)", "Test Height (mm)",
                ]
                defaults = {
                        "x pitch(mm)": "0",
                        "y pitch(mm)": "0",
                        "Number of Sensors in x": "0",
                        "Number of Sensors in y": "0",
                        "actuator speed(mm/s)": "0",
                        "speed between spaces(mm/s)": "0",
                        "max force(kg)": "0",
                        "start force(kg)": "0",
                        "start position x": "0",
                        "start position y": "0",
                        "Safe Height (mm)": "10",
                        "Test Height (mm)": "5",
                }

                self.entries: Dict[str, QLineEdit] = {}
                for idx, name in enumerate(self.setting_names):
                        r, c = divmod(idx, 2)
                        lbl = QLabel(name, self.tab_build)
                        ed = QLineEdit(self.tab_build)
                        ed.setText(defaults.get(name, ""))
                        ed.setAlignment(Qt.AlignRight)
                        ed.setMaxLength(16)
                        ed.setFixedWidth(120)
                        self.entries[name] = ed
                        grid.addWidget(lbl, r, c * 2)
                        grid.addWidget(ed, r, c * 2 + 1)

                left.addWidget(g_settings)

                # --- Checkboxes (restored core flags) ---
                g_flags = QGroupBox("Checks & Saving", self.tab_build)
                fl = QVBoxLayout(g_flags)

                # checks
                self.default_preloaded_text = "Check if Preloaded"
                self.check_preloaded = QCheckBox(self.default_preloaded_text, self.tab_build)
                self.check_preloaded.stateChanged.connect(self._on_preloaded_checked)

                self.default_shortc_text = "Check Short Circuit"
                self.check_short_circuit = QCheckBox(self.default_shortc_text, self.tab_build)
                self.check_short_circuit.stateChanged.connect(self._on_shortcircuit_checked)

                self.check_open_circuit = QCheckBox("Check Open Circuit", self.tab_build)

                # saving
                self.cb_save_raw = QCheckBox("Raw Data", self.tab_build)
                self.cb_save_filtered = QCheckBox("Filtered Data", self.tab_build)
                self.cb_same_sheet = QCheckBox("Data on Same Sheet", self.tab_build)
                self.cb_save_failed = QCheckBox("Save Failed Data", self.tab_build)

                self.cb_save_raw.setChecked(True)
                self.cb_save_filtered.setChecked(True)
                self.cb_same_sheet.setChecked(True)

                # graph flags
                self.cb_pf_enabled = QCheckBox("Pass Fail Criteria", self.tab_build)
                self.cb_graph_fxR = QCheckBox("Force x Resistance", self.tab_build)
                self.cb_graph_fxN = QCheckBox("Force x Sample Number", self.tab_build)
                self.cb_graph_rxN = QCheckBox("Resistance x Sample", self.tab_build)

                for w in (
                        self.check_preloaded, self.check_short_circuit, self.check_open_circuit,
                        self.cb_save_raw, self.cb_save_filtered, self.cb_same_sheet, self.cb_save_failed,
                        self.cb_pf_enabled, self.cb_graph_fxR, self.cb_graph_fxN, self.cb_graph_rxN
                ):
                        fl.addWidget(w)

                left.addWidget(g_flags)

                # --- Actions ---
                actions = QGroupBox("Actions", self.tab_build)
                al = QHBoxLayout(actions)

                self.btn_fill_table = QPushButton("Fill Table", self.tab_build)
                self.btn_fill_table.clicked.connect(self.fill_table)

                self.btn_add_wait = QPushButton("Add wait…", self.tab_build)
                self.btn_add_wait.clicked.connect(self.add_wait_command)

                self.btn_add_pause = QPushButton("Add pause", self.tab_build)
                self.btn_add_pause.clicked.connect(self.add_pause_command)
                
                self.btn_pf_popup = QPushButton("Pass Fail Criteria", self.tab_build)
                self.btn_pf_popup.clicked.connect(self._actions_open_pass_fail_popup)

                self.btn_gc_popup = QPushButton("Golden Curve", self.tab_build)
                self.btn_gc_popup.clicked.connect(self._actions_open_golden_curve_popup)

                self.btn_save_project = QPushButton("Save Project…", self.tab_build)
                self.btn_save_project.clicked.connect(self.save_project)

                self.btn_load_project = QPushButton("Load Project…", self.tab_build)
                self.btn_load_project.clicked.connect(self.load_from_csv)

                al.addWidget(self.btn_fill_table)
                al.addWidget(self.btn_add_wait)
                al.addWidget(self.btn_add_pause)
                al.addWidget(self.btn_pf_popup)
                al.addWidget(self.btn_gc_popup)
                al.addStretch(1)
                al.addWidget(self.btn_load_project)
                al.addWidget(self.btn_save_project)
                left.addWidget(actions)

                # right panel: table + preview log
                right = QVBoxLayout()
                outer.addLayout(right, 3)

                self.folder_display_label = QLabel("Project Folder: (not saved yet)", self.tab_build)
                right.addWidget(self.folder_display_label)

                self.grid_table = QTableWidget(self.tab_build)
                self.grid_table.setColumnCount(len(self.TABLE_HEADERS))
                self.grid_table.setHorizontalHeaderLabels(self.TABLE_HEADERS)
                self.grid_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                self.grid_table.cellDoubleClicked.connect(self._handle_grid_table_double_click)
                right.addWidget(self.grid_table, 1)

                self.build_log = QTextEdit(self.tab_build)
                self.build_log.setReadOnly(True)
                self.build_log.setPlaceholderText("Build log…")
                right.addWidget(self.build_log, 1)

        def _build_tab_run(self):
                root = QVBoxLayout(self.tab_run)

                files = QGroupBox("Run Inputs", self.tab_run)
                fl = QHBoxLayout(files)

                self.ed_plan = QLineEdit(self.tab_run)
                self.ed_plan.setPlaceholderText("Plan CSV…")
                btn_plan = QPushButton("Select Plan CSV", self.tab_run)
                btn_plan.clicked.connect(self._pick_plan)

                self.ed_out = QLineEdit(os.path.expanduser("~/Documents"), self.tab_run)
                btn_out = QPushButton("Select Output Folder", self.tab_run)
                btn_out.clicked.connect(self._pick_outdir)

                self.ed_base = QLineEdit(f"engineering_{datetime.now().strftime('%Y%m%d_%H%M%S')}", self.tab_run)
                self.ed_base.setPlaceholderText("Base name (no extension)")

                for w in (QLabel("Plan:"), self.ed_plan, btn_plan,
                                  QLabel("Output:"), self.ed_out, btn_out,
                                  QLabel("Base:"), self.ed_base):
                        fl.addWidget(w)

                root.addWidget(files)

                ctrl = QGroupBox("Controls", self.tab_run)
                cl = QHBoxLayout(ctrl)

                self.btn_start = QPushButton("Start", self.tab_run)
                self.btn_start.clicked.connect(self._start)

                self.btn_test_positions = QPushButton("Test Positions", self.tab_run)
                self.btn_test_positions.setToolTip(
                        "Move through every position in the plan like a real run, but with a "
                        "plain up/down move instead of a probe — no DAQ, no force/resistance, "
                        "no graph updates."
                )
                self.btn_test_positions.clicked.connect(self._test_positions)

                self.btn_abort = QPushButton("Abort", self.tab_run)
                self.btn_abort.clicked.connect(self._abort)
                self.btn_abort.setEnabled(False)

                cl.addStretch(1)
                cl.addWidget(self.btn_start)
                cl.addWidget(self.btn_test_positions)
                cl.addWidget(self.btn_abort)
                root.addWidget(ctrl)

                self.progress = QProgressBar(self.tab_run)
                self.progress.setRange(0, 100)
                root.addWidget(self.progress)

                self.run_log = QTextEdit(self.tab_run)
                self.run_log.setReadOnly(True)
                root.addWidget(self.run_log, 1)

                graphs = QGroupBox("Preview", self.tab_run)
                gl = QHBoxLayout(graphs)
                self.plot_force = pg.PlotWidget(title="Force")
                self.plot_res = pg.PlotWidget(title="Resistance")
                gl.addWidget(self.plot_force, 1)
                gl.addWidget(self.plot_res, 1)
                root.addWidget(graphs, 1)

        # ---------- checkbox popups ----------
        def _on_preloaded_checked(self, state: int):
                if state == Qt.Checked:
                        val, ok = QInputDialog.getDouble(
                                self, "Preload Resistance", "Enter preload resistance threshold (Ω):",
                                value=self.preload_threshold_ohm or 0.0,
                                min=0.0, decimals=2
                        )
                        if ok:
                                self.preload_threshold_ohm = float(val)
                                self.check_preloaded.setText(f"{self.default_preloaded_text} ({val:.2f} Ω)")
                        else:
                                self.check_preloaded.setChecked(False)
                else:
                        self.preload_threshold_ohm = None
                        self.check_preloaded.setText(self.default_preloaded_text)

        def _on_shortcircuit_checked(self, state: int):
                if state == Qt.Checked:
                        val, ok = QInputDialog.getDouble(
                                self, "Short-Circuit Threshold", "Enter short-circuit resistance threshold (Ω):",
                                value=self.shortcircuit_threshold_ohm or 0.0,
                                min=0.0, decimals=2
                        )
                        if ok:
                                self.shortcircuit_threshold_ohm = float(val)
                                self.check_short_circuit.setText(f"{self.default_shortc_text} ({val:.2f} Ω)")
                        else:
                                self.check_short_circuit.setChecked(False)
                else:
                        self.shortcircuit_threshold_ohm = None
                        self.check_short_circuit.setText(self.default_shortc_text)

        # ---------- table helpers ----------
        
        def _col_name(self, col: int) -> str:
            """Safely returns the header text for a column index (GRID TABLE)."""
            try:
                h = self.grid_table.horizontalHeaderItem(col)
                return (h.text() if h else "").strip()
            except Exception:
                return ""

        def _is_golden_curve_column(self, col: int) -> bool:
            return self._col_name(col).lower() == "golden curve"
        
        def _handle_grid_table_double_click(self, row: int, col: int):
                header_item = self.grid_table.horizontalHeaderItem(col)
                if header_item and header_item.text().strip().lower() == "p/f criteria":
                        dlg = PassFailCriteriaDialog(self, engineering_mode=self, apply_row=row)
                        dlg.exec_()
                elif self._is_golden_curve_column(col):
                        dlg = GoldenCurveDialog(self)
                        dlg.exec_()

                        # OPTIONAL: if you want to store the xlsx path into the cell text:
                        if dlg.xlsx_path:
                            self.grid_table.setItem(row, col, QTableWidgetItem(dlg.xlsx_path))
                                
        def _actions_open_pass_fail_popup(self):
            row = self.grid_table.currentRow()
            if row < 0:
                QMessageBox.information(self, "Select a row", "Click a row in the table first.")
                return
            dlg = PassFailCriteriaDialog(self, engineering_mode=self, apply_row=row)
            dlg.exec_()

        def _actions_open_golden_curve_popup(self):
            row = self.grid_table.currentRow()
            if row < 0:
                QMessageBox.information(self, "Select a row", "Click a row in the table first.")
                return

            dlg = GoldenCurveDialog(self)
            dlg.exec_()

            # OPTIONAL: write chosen xlsx path into the Golden Curve cell for that row
            if dlg.xlsx_path:
                gc_col = _header_index(self.grid_table, "Golden Curve")
                if gc_col >= 0:
                    self.grid_table.setItem(row, gc_col, QTableWidgetItem(dlg.xlsx_path))                        
        def apply_pf_filename(self, filename: str, only_row: Optional[int] = None):
                pf_col = _header_index(self.grid_table, "P/F Criteria")
                if pf_col < 0:
                        return
                if only_row is None:
                        for r in range(self.grid_table.rowCount()):
                                self.grid_table.setItem(r, pf_col, QTableWidgetItem(filename))
                else:
                        if 0 <= only_row < self.grid_table.rowCount():
                                self.grid_table.setItem(only_row, pf_col, QTableWidgetItem(filename))

        def add_wait_command(self):
                seconds, ok = QInputDialog.getInt(self, "Wait Duration", "Enter wait duration in seconds:", 10, 1, 10000)
                if not ok:
                        return
                row = self.grid_table.rowCount()
                self.grid_table.insertRow(row)
                self.grid_table.setItem(row, 0, QTableWidgetItem(f"wait; {seconds}"))
                for c in range(1, self.grid_table.columnCount()):
                        self.grid_table.setItem(row, c, QTableWidgetItem(""))

        def add_pause_command(self):
                row = self.grid_table.rowCount()
                self.grid_table.insertRow(row)
                self.grid_table.setItem(row, 0, QTableWidgetItem("pause"))
                for c in range(1, self.grid_table.columnCount()):
                        self.grid_table.setItem(row, c, QTableWidgetItem(""))

        # ---------- main "Fill Table" logic (ported from old GUI.py) ----------
        def fill_table(self):
                """
                Populate the grid table based on sensor counts, pitch, start position,
                and starting position dropdown.
                """
                self.grid_table.setRowCount(0)

                try:
                        x_pitch = float(self.entries["x pitch(mm)"].text())
                        y_pitch = float(self.entries["y pitch(mm)"].text())
                        sensors_x = int(float(self.entries["Number of Sensors in x"].text()))
                        sensors_y = int(float(self.entries["Number of Sensors in y"].text()))
                        start_x = float(self.entries["start position x"].text())
                        start_y = float(self.entries["start position y"].text())
                        speed_of_test = float(self.entries["actuator speed(mm/s)"].text())
                        speed_between_test = float(self.entries["speed between spaces(mm/s)"].text())
                        max_force = float(self.entries["max force(kg)"].text())
                        safe_h = float(self.entries["Safe Height (mm)"].text() or 0.0)
                        test_h = float(self.entries["Test Height (mm)"].text() or 0.0)
                except Exception as e:
                        QMessageBox.warning(self, "Input Error", f"Please enter valid numeric values.\n{e}")
                        return

                starting_position = self.dd_starting_position.currentText().strip().lower()
                total_tests = max(0, sensors_x * sensors_y)

                for test_num in range(total_tests):
                        col = test_num % sensors_x
                        row_index = test_num // sensors_x

                        if starting_position == "top right":
                                current_x = start_x - col * x_pitch
                                current_y = start_y - row_index * y_pitch
                        elif starting_position == "top left":
                                current_x = start_x + col * x_pitch
                                current_y = start_y - row_index * y_pitch
                        elif starting_position == "bottom right":
                                current_x = start_x - col * x_pitch
                                current_y = start_y + row_index * y_pitch
                        elif starting_position == "bottom left":
                                current_x = start_x + col * x_pitch
                                current_y = start_y + row_index * y_pitch
                        else:
                                current_x = start_x - col * x_pitch
                                current_y = start_y - row_index * y_pitch

                        r = self.grid_table.rowCount()
                        self.grid_table.insertRow(r)

                        # Fill
                        row_values = [
                                f"test{test_num + 1}",
                                f"{current_x:.3f}",
                                f"{current_y:.3f}",
                                str(max_force),
                                str(speed_of_test),
                                str(speed_between_test),
                                str(safe_h),
                                str(test_h),
                                "",  # Golden Curve (optional)
                                "",  # P/F Criteria
                        ]
                        for c, v in enumerate(row_values):
                                self.grid_table.setItem(r, c, QTableWidgetItem(v))

                self.build_log.append(f"[Fill] Created {total_tests} test rows.")

        # ---------- save/load project ----------
        def _collect_settings_rows(self) -> List[List[str]]:
                """
                Produce the exact "Settings" section the loader expects.

                We mirror the keys used by OperatorModePopup when building RunConfig, so your runner
                keeps working without extra translation.
                """
                rows: List[List[str]] = []
                rows.append(["Settings"])

                for k in self.setting_names:
                        rows.append([k, self.entries[k].text().strip()])

                rows.append(["Test Type", self.dd_test_type.currentText().strip()])
                rows.append(["Starting Position", self.dd_starting_position.currentText().strip()])

                # Thresholds (only meaningful if checkbox enabled; still write empty if not)
                rows.append(["Preload Resistance Threshold (Ω)", "" if self.preload_threshold_ohm is None else str(self.preload_threshold_ohm)])
                rows.append(["Short-Circuit Threshold (Ω)", "" if self.shortcircuit_threshold_ohm is None else str(self.shortcircuit_threshold_ohm)])

                # Booleans the operator runner expects
                rows.append(["Check Short Circuit", str(self.check_short_circuit.isChecked())])
                rows.append(["Check Open Circuit", str(self.check_open_circuit.isChecked())])
                rows.append(["Check if Preloaded", str(self.check_preloaded.isChecked())])

                rows.append(["Raw Data", str(self.cb_save_raw.isChecked())])
                rows.append(["Filtered Data", str(self.cb_save_filtered.isChecked())])
                rows.append(["Data on Same Sheet", str(self.cb_same_sheet.isChecked())])
                rows.append(["Save Failed Data", str(self.cb_save_failed.isChecked())])

                rows.append(["Pass Fail Criteria", str(self.cb_pf_enabled.isChecked())])
                rows.append(["Force x Resistance", str(self.cb_graph_fxR.isChecked())])
                rows.append(["Force x Sample Number", str(self.cb_graph_fxN.isChecked())])
                rows.append(["Resistance x Sample", str(self.cb_graph_rxN.isChecked())])

                return rows

        def save_project(self):
                """
                Save a project folder that contains:
                  - <name>.csv plan file
                  - pass_fail_criteria_files/<criteria>.csv for any pending criteria
                """
                base_dir = QFileDialog.getExistingDirectory(self, "Select Location to Save Project")
                if not base_dir:
                        return

                folder_name, ok = QInputDialog.getText(self, "Project Name", "Enter name for your project:", text="project")
                if not ok or not folder_name.strip():
                        return

                folder_name = folder_name.strip()
                project_folder = os.path.join(base_dir, folder_name)
                csv_path = os.path.join(project_folder, f"{folder_name}.csv")
                pf_folder = os.path.join(project_folder, "pass_fail_criteria_files")

                try:
                        os.makedirs(project_folder, exist_ok=True)
                        os.makedirs(pf_folder, exist_ok=True)
                except Exception as e:
                        QMessageBox.critical(self, "Error", f"Could not create folders:\n{e}")
                        return

                # 1) write plan CSV
                try:
                        with open(csv_path, "w", newline="") as f:
                                w = csv.writer(f)

                                # settings section
                                for row in self._collect_settings_rows():
                                        w.writerow(row)

                                # blank spacer
                                w.writerow([])

                                # table section
                                w.writerow(self.TABLE_HEADERS)
                                for r in range(self.grid_table.rowCount()):
                                        out = []
                                        for c in range(self.grid_table.columnCount()):
                                                it = self.grid_table.item(r, c)
                                                out.append("" if it is None else (it.text() or "").strip())
                                        w.writerow(out)
                except Exception as e:
                        QMessageBox.critical(self, "Save Error", f"Could not write plan CSV:\n{e}")
                        return

                # 2) write pending criteria files
                try:
                        for fname, rows in (self.pending_criteria_files or {}).items():
                                out_path = os.path.join(pf_folder, fname)
                                with open(out_path, "w", newline="") as cf:
                                        cw = csv.DictWriter(cf, fieldnames=["Type", "Value", "Max", "Min", "Other"])
                                        cw.writeheader()
                                        for r in rows:
                                                cw.writerow({
                                                        "Type": r.get("type", ""),
                                                        "Value": "" if r.get("value", None) is None else r.get("value"),
                                                        "Max": "" if r.get("max", None) is None else r.get("max"),
                                                        "Min": "" if r.get("min", None) is None else r.get("min"),
                                                        "Other": r.get("other", ""),
                                                })
                except Exception as e:
                        QMessageBox.warning(self, "Criteria Save Warning", f"Plan saved, but criteria files failed:\n{e}")

                # update UI state
                self.project_folder = project_folder
                self.pass_fail_folder = pf_folder
                self.grid_test_csv_path = csv_path
                self.folder_display_label.setText(f"Project Folder: {os.path.basename(project_folder)}")
                self.build_log.append(f"[Save] Project saved: {csv_path}")

                # also prefill Run tab
                self.ed_plan.setText(csv_path)
                self.ed_out.setText(project_folder)
                self.ed_base.setText(folder_name)

        def load_from_csv(self):
                """
                Load project settings and table from a CSV file.
                This is compatible with files saved by save_project above, and similar to old GUI.py load_from_csv.
                """
                file_path, _ = QFileDialog.getOpenFileName(self, "Open Project CSV", "", "CSV files (*.csv)")
                if not file_path:
                        return

                # avoid popups while restoring checkbox states
                self.check_preloaded.blockSignals(True)
                self.check_short_circuit.blockSignals(True)

                try:
                        with open(file_path, "r", newline="") as f:
                                rows = list(csv.reader(f))

                        # locate table header row ("Test", ...)
                        table_header_index = None
                        for i, row in enumerate(rows):
                                if row and row[0].strip().lower() == "test":
                                        table_header_index = i
                                        break
                        if table_header_index is None:
                                QMessageBox.warning(self, "Error", "CSV file missing table header row ('Test,...').")
                                return

                        # parse settings section: rows[0:table_header_index]
                        settings_dict: Dict[str, str] = {}
                        for row in rows[1:table_header_index]:
                                if len(row) >= 2 and row[0].strip():
                                        settings_dict[row[0].strip()] = row[1].strip()

                        # restore entries
                        for k in self.setting_names:
                                self.entries[k].setText(settings_dict.get(k, ""))

                        # restore dropdowns
                        if "Test Type" in settings_dict:
                                self.dd_test_type.setCurrentText(settings_dict["Test Type"])
                        if "Starting Position" in settings_dict:
                                self.dd_starting_position.setCurrentText(settings_dict["Starting Position"])

                        # thresholds
                        self.preload_threshold_ohm = None
                        if settings_dict.get("Preload Resistance Threshold (Ω)", "").strip():
                                self.preload_threshold_ohm = _safe_float(settings_dict["Preload Resistance Threshold (Ω)"], None)  # type: ignore[arg-type]
                        self.shortcircuit_threshold_ohm = None
                        if settings_dict.get("Short-Circuit Threshold (Ω)", "").strip():
                                self.shortcircuit_threshold_ohm = _safe_float(settings_dict["Short-Circuit Threshold (Ω)"], None)  # type: ignore[arg-type]

                        # checkboxes
                        def _set(cb: Optional[QCheckBox], key: str):
                                if cb is None:
                                        return
                                if key in settings_dict:
                                        cb.setChecked(_to_bool(settings_dict[key]))

                        _set(self.check_short_circuit, "Check Short Circuit")
                        _set(self.check_open_circuit, "Check Open Circuit")
                        _set(self.check_preloaded, "Check if Preloaded")
                        _set(self.cb_save_raw, "Raw Data")
                        _set(self.cb_save_filtered, "Filtered Data")
                        _set(self.cb_same_sheet, "Data on Same Sheet")
                        _set(self.cb_save_failed, "Save Failed Data")
                        _set(self.cb_pf_enabled, "Pass Fail Criteria")
                        _set(self.cb_graph_fxR, "Force x Resistance")
                        _set(self.cb_graph_fxN, "Force x Sample Number")
                        _set(self.cb_graph_rxN, "Resistance x Sample")

                        # refresh checkbox labels to include thresholds
                        if self.check_preloaded.isChecked() and self.preload_threshold_ohm is not None:
                                self.check_preloaded.setText(f"{self.default_preloaded_text} ({self.preload_threshold_ohm:.2f} Ω)")
                        else:
                                self.check_preloaded.setText(self.default_preloaded_text)

                        if self.check_short_circuit.isChecked() and self.shortcircuit_threshold_ohm is not None:
                                self.check_short_circuit.setText(f"{self.default_shortc_text} ({self.shortcircuit_threshold_ohm:.2f} Ω)")
                        else:
                                self.check_short_circuit.setText(self.default_shortc_text)

                        # restore table
                        headers = rows[table_header_index]
                        if headers and len(headers) != len(self.TABLE_HEADERS):
                                # best-effort: set to file headers
                                self.grid_table.setColumnCount(len(headers))
                                self.grid_table.setHorizontalHeaderLabels(headers)
                        else:
                                self.grid_table.setColumnCount(len(self.TABLE_HEADERS))
                                self.grid_table.setHorizontalHeaderLabels(self.TABLE_HEADERS)

                        self.grid_table.setRowCount(0)
                        for r in rows[table_header_index + 1:]:
                                if not r or not any((c or "").strip() for c in r):
                                        continue
                                rr = self.grid_table.rowCount()
                                self.grid_table.insertRow(rr)
                                for c, txt in enumerate(r[: self.grid_table.columnCount()]):
                                        self.grid_table.setItem(rr, c, QTableWidgetItem(txt))

                        # project folder labels
                        self.project_folder = os.path.dirname(file_path)
                        self.pass_fail_folder = os.path.join(self.project_folder, "pass_fail_criteria_files")
                        self.grid_test_csv_path = file_path
                        self.folder_display_label.setText(f"Project Folder: {os.path.basename(self.project_folder)}")
                        self.build_log.append(f"[Load] Loaded: {file_path}")

                        # prefill run tab
                        self.ed_plan.setText(file_path)
                        self.ed_out.setText(self.project_folder)
                        self.ed_base.setText(os.path.splitext(os.path.basename(file_path))[0])

                except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to load CSV:\n{e}")
                finally:
                        self.check_preloaded.blockSignals(False)
                        self.check_short_circuit.blockSignals(False)

        # ---------- Run tab: pickers ----------
        def _pick_plan(self):
                p, _ = QFileDialog.getOpenFileName(self, "Select Plan CSV", "", "CSV Files (*.csv);;All Files (*)")
                if p:
                        self.ed_plan.setText(p)

        def _pick_outdir(self):
                d = QFileDialog.getExistingDirectory(self, "Select Output Folder", self.ed_out.text().strip() or os.path.expanduser("~/Documents"))
                if d:
                        self.ed_out.setText(d)

        # ---------- Run tab: worker start/abort ----------
        def _start(self):
                try:
                        plan_path = self.ed_plan.text().strip()
                        if not plan_path:
                                raise ValueError("Select a plan CSV first.")
                        out_dir = self.ed_out.text().strip() or os.path.expanduser("~/Documents")
                        base = self.ed_base.text().strip() or f"engineering_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

                        steps, settings = load_grid_plan_csv(plan_path)

                        # store settings globally (runner expects it)
                        try:
                                store.set_plan(plan_path, steps, settings)  # preferred
                        except Exception:
                                store.steps = steps
                                store.settings = settings or {}

                        os.makedirs(out_dir, exist_ok=True)

                        # flags: prefer plan settings if present, but allow run tab to be "just run"
                        S = store.settings or {}
                        flags = {
                                "Raw Data": _to_bool(S.get("Raw Data", True)),
                                "Filtered Data": _to_bool(S.get("Filtered Data", True)),
                                "Data on Same Sheet": _to_bool(S.get("Data on Same Sheet", True)),
                                "Save Failed Data": _to_bool(S.get("Save Failed Data", False)),
                                "Pass Fail Criteria": _to_bool(S.get("Pass Fail Criteria", False)),
                                "Force x Resistance": _to_bool(S.get("Force x Resistance", False)),
                                "Force x Sample Number": _to_bool(S.get("Force x Sample Number", False)),
                                "Resistance x Sample": _to_bool(S.get("Resistance x Sample", False)),
                                "Check Short Circuit": _to_bool(S.get("Check Short Circuit", False)),
                                "Check Open Circuit": _to_bool(S.get("Check Open Circuit", False)),
                                "Check if Preloaded": _to_bool(S.get("Check if Preloaded", False)),
                        }

                        # writers pack (creates xlsx immediately)
                        writers = make_writers(out_dir=out_dir, flags=flags, base_name=base, job_details={})

                        # Build RunConfig from plan settings — all required fields populated.
                        def _fnum(key, cast=float, default=0.0):
                                try:
                                        v = S.get(key)
                                        if v is None or v == "":
                                                return default
                                        return cast(v)
                                except Exception:
                                        return default

                        cfg = RunConfig(
                                x_pitch=_fnum("x pitch(mm)", float, 0.0),
                                y_pitch=_fnum("y pitch(mm)", float, 0.0),
                                n_x=_fnum("Number of Sensors in x", int, 0),
                                n_y=_fnum("Number of Sensors in y", int, 0),
                                v_test=_fnum("actuator speed(mm/s)", float, 1.0),
                                v_travel=_fnum("speed between spaces(mm/s)", float, 50.0),
                                start_x=_fnum("start position x", float, 0.0),
                                start_y=_fnum("start position y", float, 0.0),
                                safe_z=_fnum("Safe Height (mm)", float, 10.0),
                                test_z=_fnum("Test Height (mm)", float, 0.0),
                                start_force=_fnum("start force(kg)", float, 0.0),
                                max_force=_fnum("max force(kg)", float, 0.0),
                                flags=flags,
                                preload_res_threshold_ohm=_fnum("Preload Resistance Threshold (Ω)", float, None),
                                short_circuit_threshold_ohm=_fnum("Short-Circuit Threshold (Ω)", float, None),
                        )
                        store.output_folder = out_dir
                        store.output_base = base
                        try:
                                store.set_run_config(cfg)
                        except Exception:
                                store.run_config = cfg

                        self.run_log.append(f"[Start] Plan: {plan_path}")
                        self.run_log.append(f"[Start] Output: {os.path.join(out_dir, base + '.xlsx')}")
                        self.run_log.append("[Start] Launching worker…")

                        if self.daq is None:
                                raise RuntimeError("DAQ not available.")
                        if hasattr(self.daq, "open") and not self._daq_ok:
                                try:
                                        self.daq.open()
                                        self._daq_ok = True
                                except Exception:
                                        pass

                        self.thread = QThread(self)
                        self.worker = TestRunnerWorker(cfg, steps, None, self.duet, self.smac, self.daq, writers)  # type: ignore
                        self.worker.moveToThread(self.thread)
                        self.thread.started.connect(self.worker.run)

                        if hasattr(self.worker, "progress"):
                                self.worker.progress.connect(self._on_progress)  # type: ignore
                        if hasattr(self.worker, "finished"):
                                self.worker.finished.connect(self._on_finished)  # type: ignore
                        if hasattr(self.worker, "error"):
                                self.worker.error.connect(self._on_error)  # type: ignore

                        self.btn_start.setEnabled(False)
                        self.btn_abort.setEnabled(True)
                        self.progress.setValue(0)

                        self.thread.start()

                except Exception as e:
                        QMessageBox.critical(self, "Start Error", str(e))

        def _test_positions(self):
                """Dry-run the plan: move through every position exactly like a
                real run (same XY travel, same safe-Z clearance, same approach
                depth), but with a plain up/down move instead of a probe move —
                no DAQ scan, no force/resistance, no graph updates."""
                try:
                        plan_path = self.ed_plan.text().strip()
                        if not plan_path:
                                raise ValueError("Select a plan CSV first.")

                        steps, settings = load_grid_plan_csv(plan_path)

                        try:
                                store.set_plan(plan_path, steps, settings)
                        except Exception:
                                store.steps = steps
                                store.settings = settings or {}

                        S = store.settings or {}

                        def _fnum(key, cast=float, default=0.0):
                                try:
                                        v = S.get(key)
                                        if v is None or v == "":
                                                return default
                                        return cast(v)
                                except Exception:
                                        return default

                        # Flags aren't used for a motion-only dry run, but RunConfig
                        # requires the field — all False since nothing is recorded.
                        flags = {k: False for k in (
                                "Raw Data", "Filtered Data", "Data on Same Sheet", "Save Failed Data",
                                "Pass Fail Criteria", "Force x Resistance", "Force x Sample Number",
                                "Resistance x Sample", "Check Short Circuit", "Check Open Circuit",
                                "Check if Preloaded",
                        )}

                        cfg = RunConfig(
                                x_pitch=_fnum("x pitch(mm)", float, 0.0),
                                y_pitch=_fnum("y pitch(mm)", float, 0.0),
                                n_x=_fnum("Number of Sensors in x", int, 0),
                                n_y=_fnum("Number of Sensors in y", int, 0),
                                v_test=_fnum("actuator speed(mm/s)", float, 1.0),
                                v_travel=_fnum("speed between spaces(mm/s)", float, 50.0),
                                start_x=_fnum("start position x", float, 0.0),
                                start_y=_fnum("start position y", float, 0.0),
                                safe_z=_fnum("Safe Height (mm)", float, 10.0),
                                test_z=_fnum("Test Height (mm)", float, 0.0),
                                start_force=_fnum("start force(kg)", float, 0.0),
                                max_force=_fnum("max force(kg)", float, 0.0),
                                flags=flags,
                                preload_res_threshold_ohm=None,
                                short_circuit_threshold_ohm=None,
                        )

                        self.run_log.append(f"[Test Positions] Plan: {plan_path}")
                        self.run_log.append(
                                f"[Test Positions] {len(steps)} position(s) — plain up/down move, "
                                "no probing, no DAQ, no graph updates."
                        )

                        pos_worker = PositionOnlyWorker(cfg, self.duet)
                        pos_runner = GridRunner(cfg=cfg, steps=steps, worker=pos_worker)

                        self.position_thread = QThread(self)
                        pos_worker.moveToThread(self.position_thread)
                        pos_runner.moveToThread(self.position_thread)
                        self.position_thread.started.connect(pos_runner.run)

                        pos_runner.progress.connect(self._on_progress)
                        pos_runner.error.connect(self._on_error)
                        pos_runner.finished.connect(self._on_positions_finished)

                        self.position_worker = pos_worker
                        self.position_runner = pos_runner

                        self.btn_start.setEnabled(False)
                        self.btn_test_positions.setEnabled(False)
                        self.btn_abort.setEnabled(True)
                        self.progress.setValue(0)

                        self.position_thread.start()

                except Exception as e:
                        QMessageBox.critical(self, "Test Positions Error", str(e))

        def _abort(self):
                if self.worker and hasattr(self.worker, "request_abort"):
                        try:
                                self.worker.request_abort()  # type: ignore
                        except Exception:
                                pass
                if self.position_runner and hasattr(self.position_runner, "request_stop"):
                        try:
                                self.position_runner.request_stop()  # type: ignore
                        except Exception:
                                pass
                self.btn_abort.setEnabled(False)

        def _on_progress(self, step_idx: int, msg: str):
                try:
                        total = len(getattr(store, "steps", None) or [])
                        pct = int(step_idx * 100 / total) if total > 0 else 0
                        self.progress.setValue(max(0, min(100, pct)))
                except Exception:
                        pass
                self.run_log.append(str(msg))

        def _on_finished(self):
                self.run_log.append("[Done] Finished.")
                self.btn_start.setEnabled(True)
                self.btn_abort.setEnabled(False)
                if self.thread:
                        self.thread.quit()
                        self.thread.wait(3000)

        def _on_positions_finished(self):
                self.run_log.append("[Test Positions] Done.")
                self.btn_start.setEnabled(True)
                self.btn_test_positions.setEnabled(True)
                self.btn_abort.setEnabled(False)
                if self.position_thread:
                        self.position_thread.quit()
                        self.position_thread.wait(3000)

        def _on_error(self, msg: str):
                self.run_log.append(f"[ERROR] {msg}")
                self.btn_start.setEnabled(True)
                self.btn_test_positions.setEnabled(True)
                self.btn_abort.setEnabled(False)

        # ---------- lifecycle ----------
        def closeEvent(self, event):
                try:
                        if self.worker and hasattr(self.worker, "request_abort"):
                                self.worker.request_abort()  # type: ignore
                except Exception:
                        pass
                try:
                        if self.thread:
                                self.thread.quit()
                                self.thread.wait(1000)
                except Exception:
                        pass
                try:
                        if self.position_runner and hasattr(self.position_runner, "request_stop"):
                                self.position_runner.request_stop()  # type: ignore
                except Exception:
                        pass
                try:
                        if self.position_thread:
                                self.position_thread.quit()
                                self.position_thread.wait(1000)
                except Exception:
                        pass
                try:
                        if self.daq is not None and hasattr(self.daq, "close"):
                                self.daq.close()
                except Exception:
                        pass
                event.accept()
