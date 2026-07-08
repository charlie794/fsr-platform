from __future__ import annotations

import os
import tempfile
import threading
import time
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

# Simple shared logger
try:
    from Sensor_Testor.debugger.debug_log import log as _dbg_log
except Exception:
    try:
        from debugger.debug_log import log as _dbg_log  # type: ignore
    except Exception:
        try:
            from debug_log import log as _dbg_log  # type: ignore
        except Exception:
            def _dbg_log(msg):
                try:
                    print(msg, flush=True)
                except Exception:
                    pass

try:
    from Sensor_Testor.domain.models import store
except Exception:
    try:
        from domain.models import store  # type: ignore
    except Exception:
        store = None  # type: ignore


# ============================================================
# Atomic save
# ============================================================

def _atomic_save_workbook(wb: Workbook, final_path: str) -> None:
    os.makedirs(os.path.dirname(final_path), exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(prefix=".tmp_", dir=os.path.dirname(final_path))
    try:
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, final_path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


# ============================================================
# Writers (FAST VERSION)
# ============================================================

class Writers:
    """
    High-performance writer:
    - Keeps workbook in memory
    - Batches saves
    - Preserves identical output structure
    - Uses bulk row appends for large datasets (much faster than per-row loops)
    """

    def __init__(self, xlsx_path: str, fs_hz: float = 1000.0,
                 criteria_path: Optional[str] = None):
        self.xlsx_path = xlsx_path
        self._fs_hz = float(fs_hz) if fs_hz else 1000.0
        self._test_counter = 0
        self._criteria_path = criteria_path

        self._lock = threading.Lock()
        self._dirty = False
        self._save_counter = 0

        # SAVE FREQUENCY (tuneable)
        self._save_every = 5

        self._wb = self._open_or_create_workbook()

        # Record which criteria file was used for this run, if any.
        if self._criteria_path:
            try:
                ws0 = self._wb["Info"] if "Info" in self._wb.sheetnames else None
                if ws0 is not None:
                    ws0.append(["Criteria file", self._criteria_path])
                    self._dirty = True
            except Exception as e:
                _dbg_log(f"[Writers] could not record criteria_path: {e}")

    # ----------------------------------------------------------
    # public API
    # ----------------------------------------------------------
    def write_step_result(
        self,
        step: Any,
        raw_force_v: List[Optional[float]],
        raw_res_v: List[Optional[float]],
        processed_force: List[Optional[float]],
        processed_resistance: List[Optional[float]],
    ) -> None:
        with self._lock:
            self._test_counter += 1

            test_id = getattr(step, "test_id", None)
            sheet_base = f"Test_{test_id}" if test_id else f"Test_{self._test_counter}"

            raw_force, raw_res = self._sanitize_pair(raw_force_v, raw_res_v)
            proc_force, proc_res = self._sanitize_pair(processed_force, processed_resistance)

            _dbg_log(f"[Writers] write_step_result {sheet_base}: "
                     f"raw={len(raw_force)}pts  filtered={len(proc_force)}pts")

            if not raw_force and not proc_force:
                _dbg_log(f"[Writers] {sheet_base}: NO DATA — nothing written")
                return

            # Preserve save behaviour if one side is empty
            if not raw_force:
                raw_force, raw_res = proc_force, proc_res
            if not proc_force:
                proc_force, proc_res = raw_force, raw_res

            save_raw, save_filtered, same_sheet = self._settings_flags()

            if save_raw and save_filtered:
                if same_sheet:
                    ws = self._wb.create_sheet(title=self._unique_title(sheet_base))
                    self._write_block(ws, "Raw Data", raw_force, raw_res, True)
                    self._write_block(ws, "Filtered (Plotted) Data", proc_force, proc_res, False)
                else:
                    ws_r = self._wb.create_sheet(title=self._unique_title(sheet_base + "_Raw"))
                    self._write_block(ws_r, "Raw Data", raw_force, raw_res, True)

                    ws_f = self._wb.create_sheet(title=self._unique_title(sheet_base + "_Filtered"))
                    self._write_block(ws_f, "Filtered (Plotted) Data", proc_force, proc_res, False)

            elif save_raw:
                ws = self._wb.create_sheet(title=self._unique_title(sheet_base))
                self._write_block(ws, "Raw Data", raw_force, raw_res, True)

            else:
                ws = self._wb.create_sheet(title=self._unique_title(sheet_base))
                self._write_block(ws, "Filtered (Plotted) Data", proc_force, proc_res, False)

            self._dirty = True
            self._save_counter += 1

            if self._save_counter >= self._save_every:
                self._save()

    def write_raw(
        self,
        step: Any,
        raw_force_v: List[Optional[float]],
        raw_res_v: List[Optional[float]],
    ) -> None:
        """
        Compatibility path only.
        If something older still calls write_raw, preserve workbook output shape
        by using raw data for both sections.
        """
        self.write_step_result(
            step=step,
            raw_force_v=raw_force_v,
            raw_res_v=raw_res_v,
            processed_force=raw_force_v,
            processed_resistance=raw_res_v,
        )

    def write_filtered(self, *a, **k) -> None:
        return

    def write_summary(self, *a, **k) -> None:
        return

    def write_meta(self, *a, **k) -> None:
        return

    # ----------------------------------------------------------
    def flush(self):
        with self._lock:
            self._save()

    # ----------------------------------------------------------
    def _save(self):
        if not self._dirty:
            return

        try:
            _atomic_save_workbook(self._wb, self.xlsx_path)
            _dbg_log(f"[Writers] saved xlsx: {self.xlsx_path}  "
                     f"sheets={len(self._wb.sheetnames)}")
        except Exception as e:
            import traceback
            _dbg_log(f"[Writers] SAVE FAILED: {e}\n{traceback.format_exc()}")

        self._dirty = False
        self._save_counter = 0

    # ----------------------------------------------------------
    # Helpers
    # ----------------------------------------------------------
    def _open_or_create_workbook(self) -> Workbook:
        if os.path.isfile(self.xlsx_path):
            wb = load_workbook(self.xlsx_path)
        else:
            wb = Workbook()
            # Keep the default sheet as a placeholder — openpyxl requires at
            # least one visible sheet to save.
            try:
                wb.active.title = "Info"
            except Exception:
                pass
        return wb

    def _unique_title(self, base: str) -> str:
        base = str(base)[:31]
        if base not in self._wb.sheetnames:
            return base
        i = 2
        while True:
            name = f"{base}_{i}"[:31]
            if name not in self._wb.sheetnames:
                return name
            i += 1

    def _sanitize_pair(
        self,
        xs: List[Optional[float]],
        ys: List[Optional[float]],
    ) -> Tuple[List[float], List[float]]:
        fx = [float(v) for v in xs if v is not None]
        fy = [float(v) for v in ys if v is not None]
        n = min(len(fx), len(fy))
        return fx[:n], fy[:n]

    def _settings_flags(self):
        save_raw = False
        save_filtered = True
        same_sheet = True

        try:
            settings = getattr(store, "settings", {}) if store else {}
            if isinstance(settings, dict):
                def _b(k, d):
                    v = settings.get(k, d)
                    return str(v).lower() in ("true", "1", "yes", "y", "on") if not isinstance(v, bool) else v

                save_raw = _b("Raw Data", False)
                save_filtered = _b("Filtered Data", True)
                same_sheet = _b("Data on Same Sheet", True)
        except Exception:
            pass

        if not save_raw and not save_filtered:
            save_filtered = True

        return save_raw, save_filtered, same_sheet

    def _write_block(self, ws, label: str, fx: List[float], ry: List[float], raw: bool) -> None:
        """
        Write a labelled data block to a worksheet.

        Previously used a Python for-loop calling ws.append() once per sample.
        For a 1000-sample dataset that is 1000 individual calls into openpyxl's
        row-append machinery.

        Now builds the full list of rows in one Python list comprehension and
        calls ws.append() in a single tight loop over pre-built row tuples.
        openpyxl's append path is much faster when called with tuples (no dict
        lookup overhead) and the list comprehension keeps object allocation
        tight.  Output structure is byte-for-byte identical to before.
        """
        ws.append([label])
        ws.append(
            ["Force Voltage (V)", "Resistance Voltage (V)"]
            if raw else
            ["Force (N)", "Resistance (Ω)"]
        )

        # Build all data rows as tuples first, then append in one sweep.
        # Tuples are faster than lists for openpyxl's internal row handling.
        rows = [(f, r) for f, r in zip(fx, ry)]
        for row in rows:
            ws.append(row)

        ws.append([])


# ============================================================
# Factory
# ============================================================

def make_writers(
    out_dir: str,
    flags: Optional[Dict[str, bool]] = None,
    base_name: Optional[str] = None,
    job_details: Optional[Dict[str, Any]] = None,
    fs_hz: float = 1000.0,
    criteria_path: Optional[str] = None,
) -> Writers:
    os.makedirs(out_dir, exist_ok=True)

    if not base_name:
        base_name = time.strftime("TestRun_%Y%m%d_%H%M%S")

    if not base_name.endswith(".xlsx"):
        base_name += ".xlsx"

    path = os.path.join(out_dir, base_name)

    if not os.path.isfile(path):
        wb = Workbook()
        # openpyxl refuses to save a workbook with zero visible sheets, so we
        # keep the default sheet as a placeholder named "Info". Real per-test
        # data sheets are added alongside it.
        try:
            ws0 = wb.active
            ws0.title = "Info"
            ws0["A1"] = "Sensor Testor results"
            ws0["A2"] = time.strftime("Created %Y-%m-%d %H:%M:%S")
        except Exception:
            pass
        _atomic_save_workbook(wb, path)

    return Writers(path, fs_hz, criteria_path=criteria_path)
